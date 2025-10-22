import os
import json
import time
import re
import csv
import io
import traceback
import base64
import requests
from flask import Flask, request, Response
from flask_cors import CORS
from openpyxl import load_workbook

app = Flask(__name__)

FRONTEND_URL = os.getenv("FRONTEND_URL", "http://localhost:5173")
CORS(app, resources={r"/api/*": {"origins": [FRONTEND_URL, "http://localhost:5173"]}})

def is_valid_email(addr):
    return isinstance(addr, str) and re.match(r"[^@]+@[^@]+\.[^@]+", addr)

def read_recipients_from_bytes(file_bytes, file_name):
    found_emails = set()
    try:
        if file_name.endswith(".csv"):
            file_text = io.StringIO(file_bytes.decode("utf-8"))
            reader = csv.reader(file_text)
            for row in reader:
                for cell_value in row:
                    cell_value = (cell_value or "").strip()
                    if is_valid_email(cell_value):
                        found_emails.add(cell_value)

        elif file_name.endswith(".xlsx") or file_name.endswith(".xls"):
            wb = load_workbook(filename=io.BytesIO(file_bytes), read_only=True)
            ws = wb.active
            if ws is None:
                raise ValueError("Worksheet is None — no active sheet found.")

            for row in ws.iter_rows():
                for cell in row:
                    cell_value = str(cell.value or "").strip()
                    if is_valid_email(cell_value):
                        found_emails.add(cell_value)
        else:
            return None
    except Exception:
        traceback.print_exc()
        return None
    return list(found_emails) if found_emails else None

@app.route("/api/send-emails-stream", methods=["POST"])
def stream_emails():
    recipients_file = request.files.get('recipientsFile')
    recipients_bytes = recipients_file.read() if recipients_file else None
    recipients_filename = recipients_file.filename if recipients_file else None
    
    attachment_file = request.files.get('attachmentFile')
    attachment_bytes = attachment_file.read() if attachment_file else None
    attachment_filename = attachment_file.filename if attachment_file else None
    
    auth_header = request.headers.get("Authorization")
    subject = request.form.get("subject")
    body_text = request.form.get("bodyText") or ""
    delay = int(request.form.get("delay", 5))

    def generate_events(token_header, sub, body, del_val, rec_bytes, rec_filename, att_bytes, att_filename):
        
        def format_event(event_type, data):
            payload = json.dumps({"type": event_type, "data": data})
            return f"data: {payload}\n\n"
            
        try:
            if not token_header:
                 yield format_event("error", {"message": "Authorization header is missing."})
                 return
            token = token_header.split(" ")[1]

            if not rec_bytes or not rec_filename:
                yield format_event("error", {"message": "Recipients file is required."})
                return
            
            recipients = read_recipients_from_bytes(rec_bytes, rec_filename)
            
            if not recipients:
                yield format_event("error", {"message": f"Could not read valid emails from {rec_filename}. Make sure it's a .xlsx or .csv file."})
                return
            
            yield format_event("log", f"Loaded {len(recipients)} recipients.")

            attachment_content_base64 = None
            if att_bytes and att_filename:
                attachment_content_base64 = base64.b64encode(att_bytes).decode('utf-8')

            endpoint = "https://graph.microsoft.com/v1.0/me/sendMail"
            headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}
            success_count, fail_count = 0, 0

            for email in recipients:
                email_msg = { 
                    "message": { 
                        "subject": sub, 
                        "body": {"contentType": "Text", "content": body_text},
                        "toRecipients": [{"emailAddress": {"address": email}}], 
                        "attachments": [] 
                    }, 
                    "saveToSentItems": "true" 
                }
                
                if attachment_content_base64:
                    email_msg["message"]["attachments"].append({ 
                        "@odata.type": "#microsoft.graph.fileAttachment", 
                        "name": att_filename, 
                        "contentBytes": attachment_content_base64 
                    })

                response = requests.post(endpoint, headers=headers, data=json.dumps(email_msg))

                if response.status_code == 202:
                    success_count += 1
                    yield format_event("progress", {"email": email, "status": "sent"})
                else:
                    fail_count += 1
                    yield format_event("progress", {"email": email, "status": "failed", "error": response.text})
                
                time.sleep(del_val)

            yield format_event("complete", { 
                "sent": success_count, 
                "failed": fail_count, 
                "message": "Process completed." 
            })

        except Exception as e:
            traceback.print_exc()
            yield format_event("error", {"message": str(e)})

    return Response(generate_events(
        auth_header, subject, body_text, delay, recipients_bytes, recipients_filename, attachment_bytes, attachment_filename
    ), mimetype='text/event-stream')